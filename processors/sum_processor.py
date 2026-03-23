# -*- coding: utf-8 -*-
"""
统计行/列总数处理器
统计指定行或列的数据总和
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .base import BaseProcessor


class SumColumnProcessor(BaseProcessor):
    """
    统计行/列总数处理器
    
    功能：统计指定行或列的数据总和
    """
    
    def __init__(self):
        """初始化处理器"""
        super(SumColumnProcessor, self).__init__()
        self.name = "统计行/列总数"
        self.description = "统计指定行或列的数据总和"
        self.params = {
            'target_row': {
                'label': '目标行',
                'type': 'row',
                'required': False,
                'default': '',
                'hint': '统计该行（与目标列二选一）'
            },
            'target_col': {
                'label': '目标列',
                'type': 'col',
                'required': False,
                'default': '',
                'hint': '统计该列（与目标行二选一）'
            }
        }
    
    def get_display_text(self, param_values=None):
        """获取带参数的显示文本"""
        if param_values:
            target_row = param_values.get('target_row', '')
            target_col = param_values.get('target_col', '')
            
            if target_row and not target_col:
                return "统计第{}行总数".format(target_row)
            elif target_col and not target_row:
                return "统计第{}列总数".format(target_col)
        return self.description
    
    def validate_params(self, param_values):
        """验证参数"""
        target_row = param_values.get('target_row', '')
        target_col = param_values.get('target_col', '')
        
        has_row = target_row and str(target_row).strip()
        has_col = target_col and str(target_col).strip()
        
        if has_row and has_col:
            return False, "目标行和目标列只能选一个"
        
        if not has_row and not has_col:
            return False, "请输入目标行或目标列"
        
        return True, ""
    
    def process(self, df, wb, sheet_name, param_values):
        """执行统计行/列总数"""
        if df.empty:
            return df, wb
        
        target_row = param_values.get('target_row', '')
        target_col = param_values.get('target_col', '')
        
        ws = wb[sheet_name]
        max_col = df.shape[1]
        max_row = df.shape[0] + 1
        
        has_row = target_row and str(target_row).strip()
        
        if has_row:
            row_num = int(target_row)
            data_row = row_num
            first_col_letter = get_column_letter(1)
            last_col_letter = get_column_letter(max_col)
            result_col = max_col + 1
            
            sum_formula = '=SUM({}{}:{}{})'.format(
                first_col_letter, data_row, last_col_letter, data_row)
            
            ws.cell(row=data_row, column=result_col, value=sum_formula)
            
            new_col_name = "行{}总数".format(row_num)
            df[new_col_name] = None
        else:
            col_num = int(target_col)
            col_letter = get_column_letter(col_num)
            data_start_row = 2
            data_end_row = max_row
            result_row = max_row + 1
            
            sum_formula = '=SUM({}{}:{}{})'.format(
                col_letter, data_start_row, col_letter, data_end_row)
            
            ws.cell(row=result_row, column=col_num, value=sum_formula)
        
        return df, wb
