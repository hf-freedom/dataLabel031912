# -*- coding: utf-8 -*-
"""
条件格式标记超出阈值处理器
将指定区域中大于/小于阈值的数据标记为红色
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.utils import get_column_letter
from .base import BaseProcessor


class HighlightProcessor(BaseProcessor):
    """
    条件格式标记超出阈值处理器
    
    功能：将指定区域中大于/小于阈值的数据标记为红色
    """
    
    def __init__(self):
        """初始化处理器"""
        super(HighlightProcessor, self).__init__()
        self.name = "条件格式标记超出阈值"
        self.description = "将指定列中从指定行到指定行中大于/小于阈值的数据标记为红色"
        self.params = {
            'target_col': {
                'label': '目标列',
                'type': 'col',
                'required': True,
                'default': 1,
                'hint': '要检查数据的列号（如第1列填1）'
            },
            'start_row': {
                'label': '开始行',
                'type': 'row',
                'required': True,
                'default': 2,
                'hint': '从该行开始检查（第1行是表头）'
            },
            'end_row': {
                'label': '结束行',
                'type': 'row',
                'required': False,
                'default': '',
                'hint': '检查到该行（留空则到最后一行）'
            },
            'operator': {
                'label': '比较方式',
                'type': 'text',
                'required': True,
                'default': '大于',
                'hint': '大于 或 小于'
            },
            'threshold': {
                'label': '阈值',
                'type': 'text',
                'hint': '用于比较的阈值（数字）'
            }
        }
    
    def get_display_text(self, param_values=None):
        """获取带参数的显示文本"""
        if param_values:
            target_col = param_values.get('target_col', '1')
            start_row = param_values.get('start_row', '')
            end_row = param_values.get('end_row', '')
            operator = param_values.get('operator', '大于')
            threshold = param_values.get('threshold', '')
            
            if start_row and threshold and target_col:
                end_text = end_row if end_row else '最后'
                return "将第{}列第{}行到第{}行中{}{}的数据标记为红色".format(target_col, start_row, end_text, operator, threshold)
        return self.description
    
    def validate_params(self, param_values):
        """验证参数"""
        target_col = param_values.get('target_col', '')
        start_row = param_values.get('start_row', '')
        operator = param_values.get('operator', '')
        threshold = param_values.get('threshold', '')
        
        if not target_col or not str(target_col).strip():
            return False, "请输入目标列"
        
        try:
            target_col = int(target_col)
        except (ValueError, TypeError):
            return False, "目标列必须是数字"
        
        if target_col < 1:
            return False, "目标列必须大于等于1"
        
        if not start_row or not str(start_row).strip():
            return False, "请输入开始行"
        
        try:
            start_row = int(start_row)
        except (ValueError, TypeError):
            return False, "开始行必须是数字"
        
        if start_row < 2:
            return False, "开始行必须大于等于2（第1行是表头）"
        
        end_row = param_values.get('end_row', '')
        if end_row and str(end_row).strip():
            try:
                end_row = int(end_row)
                if end_row < start_row:
                    return False, "结束行必须大于等于开始行"
            except (ValueError, TypeError):
                return False, "结束行必须是数字"
        
        if not operator:
            return False, "比较方式请输入：大于 或 小于"
        
        if not threshold or not str(threshold).strip():
            return False, "请输入阈值"
        
        try:
            float(threshold)
        except (ValueError, TypeError):
            return False, "阈值必须是数字"
        
        return True, ""
    
    def process(self, df, wb, sheet_name, param_values):
        """执行条件格式标记"""
        if df.empty:
            return df, wb
        
        target_col = int(self.get_param_value(param_values, 'target_col', 1))
        start_row = int(self.get_param_value(param_values, 'start_row', 2))
        end_row = self.get_param_value(param_values, 'end_row', '')
        operator = self.get_param_value(param_values, 'operator', '大于')
        threshold = float(self.get_param_value(param_values, 'threshold', 0))
        
        ws = wb[sheet_name]
        
        max_col = df.shape[1]
        max_row = df.shape[0] + 1
        
        # 确保目标列在有效范围内
        if target_col < 1 or target_col > max_col:
            return df, wb
        
        if end_row and str(end_row).strip():
            end_row = int(end_row)
        else:
            end_row = max_row
        
        start_row = max(start_row, 2)
        end_row = min(end_row, max_row)
        
        if start_row > end_row:
            return df, wb
        
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True)
        
        target_col_letter = get_column_letter(target_col)
        
        # 条件格式只应用于目标列
        apply_range = "{}{}:{}{}".format(target_col_letter, start_row, target_col_letter, end_row)
        
        op_type = '>'
        if '>=' in operator or operator in ['>=', u'>=', u'大于等于']:
            op_type = '>='
        elif '>' in operator or operator in ['>', u'>', u'大于']:
            op_type = '>'
        elif '<=' in operator or operator in ['<=', u'<=', u'小于等于']:
            op_type = '<='
        elif '<' in operator or operator in ['<', u'<', u'小于']:
            op_type = '<'
        elif '=' in operator or operator in ['=', u'=', u'等于']:
            op_type = '='
        
        formula = "AND(ISNUMBER({}{}), {}{}{}{})".format(target_col_letter, start_row, target_col_letter, start_row, op_type, threshold)
        
        try:
            dxf = DifferentialStyle(fill=red_fill, font=white_font)
            rule = Rule(type='expression', dxf=dxf, stopIfTrue=False)
            rule.formula = [formula]
            ws.conditional_formatting.add(apply_range, rule)
        except Exception:
            pass
        
        # 只遍历目标列
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=target_col)
            cell_value = cell.value
            
            if cell_value is None:
                continue
            
            try:
                cell_numeric = float(cell_value)
            except (ValueError, TypeError):
                continue
            
            should_highlight = False
            if op_type == '>' and cell_numeric > threshold:
                should_highlight = True
            elif op_type == '<' and cell_numeric < threshold:
                should_highlight = True
            elif op_type == '=' and cell_numeric == threshold:
                should_highlight = True
            elif op_type == '>=' and cell_numeric >= threshold:
                should_highlight = True
            elif op_type == '<=' and cell_numeric <= threshold:
                should_highlight = True
            
            if should_highlight:
                cell.fill = red_fill
                cell.font = white_font
        
        return df, wb
