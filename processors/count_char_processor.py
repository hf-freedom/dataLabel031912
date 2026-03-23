# -*- coding: utf-8 -*-
"""
统计字符出现次数处理器
统计指定字符在某行或某列出现的次数
"""

from typing import Tuple, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .base import BaseProcessor


class CountCharProcessor(BaseProcessor):
    """
    统计字符出现次数处理器
    
    功能：统计指定字符在某行或某列出现的次数
    """
    
    def __init__(self):
        """初始化处理器"""
        super().__init__()
        self.name = "统计字符出现次数"
        self.description = "统计指定字符在某行或某列出现的次数"
        self.params = {
            'target_string': {
                'label': '统计字符',
                'type': 'text',
                'required': True,
                'default': 'aa',
                'hint': '要统计的字符'
            },
            'target_row': {
                'label': '目标行',
                'type': 'row',
                'required': False,
                'default': '',
                'hint': '统计该行（与目标列二选一）'
            },
            'target_col': {
                'label': '目标列',
                'type': 'col',
                'required': False,
                'default': '',
                'hint': '统计该列（与目标行二选一）'
            }
        }
    
    def get_display_text(self, param_values: Dict[str, Any] = None) -> str:
        """获取带参数的显示文本"""
        if param_values:
            target_string = param_values.get('target_string', '')
            target_row = param_values.get('target_row', '')
            target_col = param_values.get('target_col', '')
            
            if target_row and not target_col:
                return f"统计'{target_string}'在第{target_row}行出现的次数"
            elif target_col and not target_row:
                return f"统计'{target_string}'在第{target_col}列出现的次数"
        return self.description
    
    def validate_params(self, param_values: Dict[str, Any]) -> Tuple[bool, str]:
        """验证参数"""
        target_row = param_values.get('target_row', '')
        target_col = param_values.get('target_col', '')
        target_string = param_values.get('target_string', '')
        
        if not target_string:
            return False, "请输入要统计的字符"
        
        has_row = target_row and str(target_row).strip()
        has_col = target_col and str(target_col).strip()
        
        if has_row and has_col:
            return False, "目标行和目标列只能选一个"
        
        if not has_row and not has_col:
            return False, "请输入目标行或目标列"
        
        return True, ""
    
    def process(self, df: pd.DataFrame, wb: Workbook, sheet_name: str,
                param_values: Dict[str, Any]) -> Tuple[pd.DataFrame, Workbook]:
        """执行统计字符出现次数"""
        if df.empty:
            return df, wb
        
        target_string = self.get_param_value(param_values, 'target_string', 'aa')
        target_row = param_values.get('target_row', '')
        target_col = param_values.get('target_col', '')
        
        ws = wb[sheet_name]
        max_col = df.shape[1]
        max_row = df.shape[0] + 1
        
        has_row = target_row and str(target_row).strip()
        
        if has_row:
            row_num = int(target_row)
            data_row = row_num + 1
            first_col_letter = get_column_letter(1)
            last_col_letter = get_column_letter(max_col)
            result_col = max_col + 1
            
            countif_formula = f'=COUNTIF({first_col_letter}{data_row}:{last_col_letter}{data_row},"{target_string}")'
            
            ws.cell(row=data_row, column=result_col, value=countif_formula)
            
            new_col_name = f"{target_string}计数_行{row_num}"
            df[new_col_name] = None
        else:
            col_num = int(target_col)
            col_letter = get_column_letter(col_num)
            data_start_row = 2
            data_end_row = max_row
            result_row = max_row + 1
            
            countif_formula = f'=COUNTIF({col_letter}{data_start_row}:{col_letter}{data_end_row},"{target_string}")'
            
            ws.cell(row=result_row, column=col_num, value=countif_formula)
        
        return df, wb
