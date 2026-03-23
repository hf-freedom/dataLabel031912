# -*- coding: utf-8 -*-
"""
条件格式标记超出阈值处理器
将指定区域中大于/小于阈值的数据标记为红色
"""

from typing import Tuple, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from .base import BaseProcessor


class HighlightThresholdProcessor(BaseProcessor):
    """
    条件格式标记超出阈值处理器
    
    功能：将指定区域中大于/小于阈值的数据标记为红色
    使用条件格式，新数据也能自动变色
    """
    
    def __init__(self):
        super().__init__()
        self.name = "条件格式标记超出阈值"
        self.description = "将指定区域中大于/小于阈值的数据标记为红色"
        self.params = {
            'start_row': {
                'label': '开始行',
                'type': 'row',
                'required': True,
                'default': 2,
                'hint': '标记起始行'
            },
            'end_row': {
                'label': '结束行',
                'type': 'row',
                'required': True,
                'default': 10,
                'hint': '标记结束行'
            },
            'compare_type': {
                'label': '比较方式',
                'type': 'text',
                'required': True,
                'default': '大于',
                'hint': '大于或小于'
            },
            'threshold': {
                'label': '阈值',
                'type': 'text',
                'required': True,
                'default': '100',
                'hint': '比较的阈值'
            },
            'target_col': {
                'label': '目标列',
                'type': 'col',
                'required': True,
                'default': 1,
                'hint': '要标记的列'
            }
        }
    
    def get_display_text(self, param_values: Dict[str, Any] = None) -> str:
        if param_values:
            start_row = param_values.get('start_row', '')
            end_row = param_values.get('end_row', '')
            compare_type = param_values.get('compare_type', '')
            threshold = param_values.get('threshold', '')
            if start_row and end_row and compare_type and threshold:
                return f"将第{start_row}行到第{end_row}行中{compare_type}{threshold}的数据标记为红色"
        return self.description
    
    def validate_params(self, param_values: Dict[str, Any]) -> Tuple[bool, str]:
        start_row = param_values.get('start_row', '')
        end_row = param_values.get('end_row', '')
        compare_type = param_values.get('compare_type', '')
        threshold = param_values.get('threshold', '')
        target_col = param_values.get('target_col', '')
        
        if not start_row:
            return False, "请输入开始行"
        if not end_row:
            return False, "请输入结束行"
        if not compare_type:
            return False, "请输入比较方式"
        if not threshold:
            return False, "请输入阈值"
        if not target_col:
            return False, "请输入目标列"
        
        try:
            start = int(start_row)
            end = int(end_row)
            if start > end:
                return False, "开始行不能大于结束行"
            if start < 1:
                return False, "开始行必须大于0"
        except ValueError:
            return False, "行号必须是数字"
        
        if compare_type not in ['大于', '小于']:
            return False, "比较方式必须是'大于'或'小于'"
        
        try:
            float(threshold)
        except ValueError:
            return False, "阈值必须是数字"
        
        return True, ""
    
    def process(self, df: pd.DataFrame, wb: Workbook, sheet_name: str,
                param_values: Dict[str, Any]) -> Tuple[pd.DataFrame, Workbook]:
        if df.empty:
            return df, wb
        
        start_row = self.get_param_value(param_values, 'start_row', 2)
        end_row = self.get_param_value(param_values, 'end_row', 10)
        compare_type = self.get_param_value(param_values, 'compare_type', '大于')
        threshold = self.get_param_value(param_values, 'threshold', '100')
        target_col = self.get_param_value(param_values, 'target_col', 1)
        
        ws = wb[sheet_name]
        
        col_letter = get_column_letter(target_col)
        cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
        
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        
        threshold_value = float(threshold)
        
        if compare_type == '大于':
            rule = CellIsRule(
                operator='greaterThan',
                formula=[str(threshold_value)],
                fill=red_fill,
                font=white_font
            )
        else:
            rule = CellIsRule(
                operator='lessThan',
                formula=[str(threshold_value)],
                fill=red_fill,
                font=white_font
            )
        
        ws.conditional_formatting.add(cell_range, rule)
        
        return df, wb
