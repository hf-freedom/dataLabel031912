# -*- coding: utf-8 -*-
"""
测试条件格式功能
"""

import pandas as pd
from openpyxl import load_workbook
from processors.highlight_processor import HighlightProcessor

# 创建测试数据
test_data = {
    'ID': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Eve', 'Frank', 'Grace', 'Henry', 'Ivy', 'Jack'],
    'Score': [85, 92, 78, 95, 88, 72, 91, 83, 76, 89],
    'Points': [120, 150, 98, 145, 132, 88, 148, 125, 95, 138]
}

# 创建DataFrame
df = pd.DataFrame(test_data)

# 保存为Excel文件
test_file_path = 'test_conditional.xlsx'
df.to_excel(test_file_path, index=False, sheet_name='Sheet1')

print("Test file created: {}".format(test_file_path))
print("\nTest data:")
print(df)

# 加载工作簿
wb = load_workbook(test_file_path)
sheet_name = wb.sheetnames[0]

# 创建处理器并应用条件格式
processor = HighlightProcessor()

# 测试参数 - 标记Score列大于90的单元格
param_values = {
    'start_row': 2,
    'end_row': 11,
    'condition': '大于',
    'threshold': 90,
    'target_col': 3  # Score列
}

print("\n" + "="*50)
print("Applying conditional formatting...")
print("="*50)

# 验证参数
valid, error_msg = processor.validate_params(param_values)
print("Parameter validation: {}, {}".format(valid, error_msg if error_msg else 'passed'))

if valid:
    df_result, wb_result = processor.process(df.copy(), wb, sheet_name, param_values)
    
    # 保存结果
    output_path = 'test_conditional_result.xlsx'
    wb_result.save(output_path)
    print("\nResult saved: {}".format(output_path))
    print("\nInstructions:")
    print("1. Open test_conditional_result.xlsx in Excel")
    print("2. Check column C (Score), values > 90 should be RED")
    print("3. Try changing a value (e.g., change 85 to 95)")
    print("4. The cell should automatically turn RED")

wb.close()

# 验证条件格式是否正确应用
print("\n" + "="*50)
print("Verifying conditional formatting...")
print("="*50)

wb_check = load_workbook(output_path)
ws_check = wb_check.active

# 检查条件格式规则
print("\nConditional formatting rules:")
for rule in ws_check.conditional_formatting:
    print("Range: {}".format(rule))
    for r in ws_check.conditional_formatting[rule]:
        print("  Rule type: {}".format(type(r).__name__))
        if hasattr(r, 'operator'):
            print("  Operator: {}".format(r.operator))
        if hasattr(r, 'formula'):
            print("  Formula: {}".format(r.formula))
        if hasattr(r, 'fill'):
            print("  Fill color: {}".format(r.fill.start_color.rgb if r.fill else 'None'))

wb_check.close()

print("\n" + "="*50)
print("Test completed!")
print("="*50)
