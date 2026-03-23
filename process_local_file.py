# -*- coding: utf-8 -*-
"""
处理本地test.xlsx文件 - 第4列大于3的标记为红色
"""

import pandas as pd
from openpyxl import load_workbook
from processors.highlight_processor import HighlightProcessor

# 本地文件路径
local_file = r'C:\Users\12824\Desktop\dataLabel\0319\p12\test.xlsx'
output_file = r'C:\Users\12824\Desktop\dataLabel\0319\p12\test_highlighted.xlsx'

print("Processing local file: {}".format(local_file))
print("="*50)

# 读取文件
df = pd.read_excel(local_file)
print("File content:")
print(df)
print("\nColumn 4 (数量) values: {}".format(df.iloc[:, 3].tolist()))
print("Values > 3: {}".format(df[df.iloc[:, 3] > 3].iloc[:, 3].tolist()))

# 加载工作簿
wb = load_workbook(local_file)
sheet_name = wb.sheetnames[0]

# 创建处理器
processor = HighlightProcessor()

# 参数 - 第4列（数量列），大于3的标记为红色
param_values = {
    'start_row': 2,
    'end_row': 8,  # 根据实际数据行数调整
    'condition': '大于',
    'threshold': 3,
    'target_col': 4  # 第4列
}

print("\n" + "="*50)
print("Applying conditional formatting...")
print("Column: 4 (数量)")
print("Condition: > 3")
print("="*50)

# 验证参数
valid, error_msg = processor.validate_params(param_values)
print("Validation: {}".format(error_msg if error_msg else 'passed'))

if valid:
    df_result, wb_result = processor.process(df.copy(), wb, sheet_name, param_values)
    
    # 保存结果
    wb_result.save(output_file)
    print("\nResult saved: {}".format(output_file))
    print("\nInstructions:")
    print("1. Open test_highlighted.xlsx in Excel")
    print("2. Column D (数量) values > 3 should be RED")
    print("3. Try changing a value:")
    print("   - Change 5 to 2 -> should turn NORMAL")
    print("   - Change 2 to 5 -> should turn RED automatically")

wb.close()

print("\n" + "="*50)
print("Done!")
print("="*50)
