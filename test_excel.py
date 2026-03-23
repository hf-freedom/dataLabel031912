# -*- coding: utf-8 -*-
"""
Test script: Create test Excel file and test processor functions
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# Create test data
test_data = {
    'ID': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
    'Name': ['Alice', 'Bob', 'Charlie', 'David', 'Eve', 'Frank', 'Grace', 'Henry', 'Ivy', 'Jack'],
    'Score': [85, 92, 78, 95, 88, 72, 91, 83, 76, 89],
    'Points': [120, 150, 98, 145, 132, 88, 148, 125, 95, 138]
}

# Create DataFrame
df = pd.DataFrame(test_data)

# Save to Excel file
test_file_path = os.path.join(os.path.dirname(__file__), 'test_data.xlsx')
df.to_excel(test_file_path, index=False, sheet_name='Sheet1')

print("Test file created: {}".format(test_file_path))
print("\nTest data:")
print(df)

# Test column sort processor
print("\n" + "="*50)
print("Test Column Sort Processor")
print("="*50)

from processors.column_sort_processor import ColumnSortProcessor

sort_processor = ColumnSortProcessor()
print("Processor name: {}".format(sort_processor.name))
print("Processor description: {}".format(sort_processor.description))
print("Params: {}".format(sort_processor.params))

# Load workbook for testing
from openpyxl import load_workbook

wb = load_workbook(test_file_path)
sheet_name = wb.sheetnames[0]

# Test parameters
param_values = {
    'sort_col': 3,  # Sort by Score column
    'start_row': 2,
    'end_row': 11,
    'sort_order': '降序'
}

# Validate parameters
valid, error_msg = sort_processor.validate_params(param_values)
print("\nParameter validation: {}, {}".format(valid, error_msg if error_msg else 'passed'))

if valid:
    df_result, wb_result = sort_processor.process(df.copy(), wb, sheet_name, param_values)
    print("\nSorted data:")
    print(df_result)
    
    # Save result
    output_path = os.path.join(os.path.dirname(__file__), 'test_sort_result.xlsx')
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_result.to_excel(writer, sheet_name=sheet_name, index=False)
    print("\nSort result saved: {}".format(output_path))

wb.close()

# Test highlight processor
print("\n" + "="*50)
print("Test Highlight Processor")
print("="*50)

from processors.highlight_processor import HighlightProcessor

highlight_processor = HighlightProcessor()
print("Processor name: {}".format(highlight_processor.name))
print("Processor description: {}".format(highlight_processor.description))
print("Params: {}".format(highlight_processor.params))

# Reload workbook
wb = load_workbook(test_file_path)

# Test parameters
param_values = {
    'start_row': 2,
    'end_row': 11,
    'condition': '大于',
    'threshold': 90,
    'target_col': 3  # Highlight Score column
}

# Validate parameters
valid, error_msg = highlight_processor.validate_params(param_values)
print("\nParameter validation: {}, {}".format(valid, error_msg if error_msg else 'passed'))

if valid:
    df_result, wb_result = highlight_processor.process(df.copy(), wb, sheet_name, param_values)
    
    # Save result
    output_path = os.path.join(os.path.dirname(__file__), 'test_highlight_result.xlsx')
    wb_result.save(output_path)
    print("\nHighlight result saved: {}".format(output_path))
    print("Please open the file to see red highlighting effect")

wb.close()

print("\n" + "="*50)
print("All tests completed!")
print("="*50)
