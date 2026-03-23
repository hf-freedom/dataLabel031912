# -*- coding: utf-8 -*-
"""
验证高亮标记是否正确应用
"""

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 加载高亮结果文件
wb = load_workbook('test_highlight_result.xlsx')
ws = wb.active

print("Checking highlight in test_highlight_result.xlsx")
print("="*50)

# 检查第3列（Score列）从第2行到第11行
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
highlighted_cells = []

for row_idx in range(2, 12):  # 第2行到第11行
    cell = ws.cell(row=row_idx, column=3)  # 第3列（Score列）
    cell_value = cell.value
    cell_fill = cell.fill
    
    # 检查是否是红色填充
    is_red = (cell_fill.start_color.rgb == 'FFFF0000' or 
              cell_fill.start_color.rgb == 'FF0000' or
              cell_fill.fgColor.rgb == 'FFFF0000' or
              cell_fill.fgColor.rgb == 'FF0000')
    
    if is_red:
        highlighted_cells.append((row_idx, cell_value))
        print("Row {} (Value: {}) - RED".format(row_idx, cell_value))
    else:
        print("Row {} (Value: {}) - normal".format(row_idx, cell_value))

print("\n" + "="*50)
print("Summary:")
print("Total highlighted cells: {}".format(len(highlighted_cells)))
print("Highlighted values (>90): {}".format([cell[1] for cell in highlighted_cells]))

# 验证哪些值应该被高亮（大于90）
expected_values = [92, 95, 91]
actual_values = [cell[1] for cell in highlighted_cells]

print("\nExpected values > 90: {}".format(expected_values))
print("Actual highlighted: {}".format(actual_values))

if set(actual_values) == set(expected_values):
    print("\n✓ SUCCESS: All values > 90 are highlighted correctly!")
else:
    print("\n✗ MISMATCH: Some values are not highlighted correctly")

wb.close()
