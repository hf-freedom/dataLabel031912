# -*- coding: utf-8 -*-
"""
Excel处理工具主程序
提供图形化界面，支持多种Excel处理功能
"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

from processors.base import BaseProcessor
from processors import (
    SortFirstColumnProcessor,
    CountCharProcessor,
    SumColumnProcessor,
    ColumnSortProcessor,
    HighlightProcessor
)

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
    HAS_DND = True
except ImportError:
    HAS_DND = False


class ExcelProcessorApp:
    """
    Excel处理工具主应用类
    
    界面结构：
    1. 功能选择区 - 下拉框选择功能
    2. 参数设置区 - 根据功能动态生成参数输入框
    3. 文件选择区 - 选择或拖入Excel文件
    4. 执行按钮
    """
    
    def __init__(self, root):
        """初始化应用"""
        self.root = root
        self.root.title("Excel处理工具")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        self.processors = {}
        self.param_vars = {}
        self.param_entries = {}
        
        self._register_processors()
        
        self.file_path = None
        
        self._setup_ui()
    
    def _register_processors(self):
        """注册所有处理器"""
        processor_classes = [
            SortFirstColumnProcessor,
            CountCharProcessor,
            SumColumnProcessor,
            ColumnSortProcessor,
            HighlightProcessor
        ]

        for proc_class in processor_classes:
            proc = proc_class()
            self.processors[proc.name] = proc
    
    def _setup_ui(self):
        """设置用户界面"""
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        self._setup_function_area(main_frame)
        
        self._setup_param_area(main_frame)
        
        self._setup_file_area(main_frame)
        
        self._setup_execute_button(main_frame)
        
        self._setup_status_area(main_frame)
    
    def _setup_function_area(self, parent):
        """设置功能选择区域"""
        func_frame = ttk.LabelFrame(parent, text="功能选择", padding="10")
        func_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(func_frame, text="选择功能：").pack(side=tk.LEFT, padx=(0, 10))
        
        self.func_var = tk.StringVar()
        processor_names = list(self.processors.keys())
        
        self.func_combo = ttk.Combobox(
            func_frame, 
            textvariable=self.func_var,
            values=processor_names,
            state="readonly",
            width=30
        )
        self.func_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        if processor_names:
            self.func_combo.current(0)
        
        self.func_combo.bind("<<ComboboxSelected>>", lambda e: self._on_function_changed())
    
    def _setup_param_area(self, parent):
        """设置参数设置区域"""
        self.param_frame = ttk.LabelFrame(parent, text="参数设置", padding="10")
        self.param_frame.pack(fill=tk.X, pady=(0, 10))
        
        self._build_param_inputs()
    
    def _build_param_inputs(self):
        """根据当前选择的处理器构建参数输入框"""
        for widget in self.param_frame.winfo_children():
            widget.destroy()
        
        self.param_vars.clear()
        self.param_entries.clear()
        
        func_name = self.func_var.get()
        processor = self.processors.get(func_name)
        
        if not processor or not processor.params:
            ttk.Label(self.param_frame, text="该功能无需参数", foreground="gray").pack()
            return
        
        for param_name, param_def in processor.params.items():
            param_frame = ttk.Frame(self.param_frame)
            param_frame.pack(fill=tk.X, pady=(0, 5))
            
            label = ttk.Label(param_frame, text=f"{param_def['label']}：")
            label.pack(side=tk.LEFT, padx=(0, 10))
            
            var = tk.StringVar()
            var.trace_add("write", lambda *args: self._update_hint())
            self.param_vars[param_name] = var
            
            entry = ttk.Entry(param_frame, textvariable=var, width=15)
            entry.pack(side=tk.LEFT)
            
            hint = param_def.get('hint', '')
            if hint:
                ttk.Label(param_frame, text=hint, foreground="gray").pack(side=tk.LEFT, padx=(10, 0))
        
        self.hint_label = ttk.Label(
            self.param_frame, 
            text="", 
            foreground="blue", 
            font=("", 10, "bold")
        )
        self.hint_label.pack(fill=tk.X, pady=(10, 0))
        
        self._update_hint()
    
    def _on_function_changed(self):
        """功能选择变更时重新构建参数输入框"""
        self._build_param_inputs()
    
    def _update_hint(self):
        """更新提示语"""
        func_name = self.func_var.get()
        processor = self.processors.get(func_name)
        
        if not processor:
            return
        
        param_values = {}
        for param_name, var in self.param_vars.items():
            value = var.get().strip()
            if value.isdigit():
                param_values[param_name] = int(value)
            else:
                param_values[param_name] = value
        
        display_text = processor.get_display_text(param_values)
        self.hint_label.config(text=display_text)
    
    def _setup_file_area(self, parent):
        """设置文件选择区域"""
        file_frame = ttk.LabelFrame(parent, text="文件选择", padding="10")
        file_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        btn_frame = ttk.Frame(file_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Button(btn_frame, text="选择文件", command=self._select_file).pack(side=tk.LEFT, padx=(0, 10))
        
        self.file_label = ttk.Label(btn_frame, text="未选择文件", foreground="gray")
        self.file_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        self.drop_frame = tk.Frame(file_frame, relief="groove", borderwidth=2, bg="#f0f0f0")
        self.drop_frame.pack(fill=tk.BOTH, expand=True)
        
        drop_text = "将Excel文件拖拽到此处\n或点击此处选择文件"
        if not HAS_DND:
            drop_text = "点击此处选择Excel文件\n（安装tkinterdnd2可启用拖拽功能）"
        
        self.drop_label = tk.Label(
            self.drop_frame, 
            text=drop_text,
            anchor="center",
            justify="center",
            bg="#f0f0f0",
            fg="#666666"
        )
        self.drop_label.pack(fill=tk.BOTH, expand=True, pady=30)
        
        self._setup_drag_drop()
    
    def _setup_drag_drop(self):
        """设置拖拽功能"""
        if HAS_DND:
            self.drop_frame.drop_target_register(DND_FILES)
            self.drop_frame.dnd_bind('<<Drop>>', self._on_drop)
            self.drop_frame.dnd_bind('<<DragEnter>>', lambda e: self._on_drag_enter(e))
            self.drop_frame.dnd_bind('<<DragLeave>>', lambda e: self._on_drag_leave(e))
        
        self.drop_frame.bind("<Button-1>", lambda e: self._select_file())
        self.drop_label.bind("<Button-1>", lambda e: self._select_file())
    
    def _on_drop(self, event):
        """文件拖放事件"""
        files = event.data
        if files.startswith('{'):
            files = files[1:-1]
        file_list = files.split()
        if file_list:
            file_path = file_list[0]
            if file_path.lower().endswith(('.xlsx', '.xls')):
                self._set_file(file_path)
            else:
                messagebox.showwarning("提示", "请拖入Excel文件（.xlsx或.xls）")
        self._on_drag_leave(None)
    
    def _on_drag_enter(self, event):
        """拖拽进入"""
        self.drop_frame.config(bg="#d0e0f0")
        self.drop_label.config(bg="#d0e0f0", fg="#333333")
    
    def _on_drag_leave(self, event):
        """拖拽离开"""
        self.drop_frame.config(bg="#f0f0f0")
        self.drop_label.config(bg="#f0f0f0", fg="#666666")
    
    def _set_file(self, file_path):
        """设置当前选择的文件"""
        self.file_path = os.path.normpath(file_path)
        self.file_label.config(text=os.path.basename(self.file_path), foreground="black")
        self.drop_label.config(text=f"已选择：\n{os.path.basename(self.file_path)}")
        self.status_var.set(f"已加载文件：{os.path.basename(self.file_path)}")
    
    def _setup_execute_button(self, parent):
        """设置执行按钮"""
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=tk.X)
        
        self.execute_btn = ttk.Button(btn_frame, text="执行处理", command=self._execute)
        self.execute_btn.pack(fill=tk.X, ipady=10)
    
    def _setup_status_area(self, parent):
        """设置状态显示区域"""
        self.status_var = tk.StringVar(value="就绪")
        status_label = ttk.Label(
            parent, 
            textvariable=self.status_var,
            relief="sunken",
            anchor="w",
            padding=(5, 2)
        )
        status_label.pack(fill=tk.X, pady=(10, 0))
    
    def _select_file(self):
        """打开文件选择对话框"""
        file_path = filedialog.askopenfilename(
            title="选择Excel文件",
            filetypes=[
                ("Excel文件", "*.xlsx *.xls"),
                ("所有文件", "*.*")
            ]
        )
        if file_path:
            self._set_file(file_path)
    
    def _get_param_values(self):
        """获取参数值字典"""
        param_values = {}
        for param_name, var in self.param_vars.items():
            value = var.get().strip()
            if value.isdigit():
                param_values[param_name] = int(value)
            else:
                param_values[param_name] = value
        return param_values
    
    def _execute(self):
        """执行处理操作"""
        if not self.file_path:
            messagebox.showerror("错误", "请先选择Excel文件")
            return
        
        func_name = self.func_var.get()
        processor = self.processors.get(func_name)
        
        if not processor:
            messagebox.showerror("错误", "请选择有效的处理功能")
            return
        
        param_values = self._get_param_values()
        
        valid, error_msg = processor.validate_params(param_values)
        if not valid:
            messagebox.showerror("输入错误", error_msg)
            return
        
        self.status_var.set("正在处理...")
        self.root.update()
        
        try:
            self.file_path = os.path.normpath(self.file_path)
            
            df = pd.read_excel(self.file_path)
            
            wb = load_workbook(self.file_path)
            sheet_name = wb.sheetnames[0]
            
            df_result, wb_result = processor.process(df, wb, sheet_name, param_values)
            
            output_path = self._get_output_path()
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df_result.to_excel(writer, sheet_name=sheet_name, index=False)
            
            wb_save = load_workbook(output_path)
            ws_save = wb_save[sheet_name]
            ws_orig = wb_result[sheet_name]
            
            for row in ws_orig.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        ws_save[cell.coordinate].value = cell.value
            
            wb_save.save(output_path)
            wb_save.close()
            wb.close()
            
            self.status_var.set(f"处理完成！文件已保存至：{os.path.basename(output_path)}")
            messagebox.showinfo("成功", f"处理完成！\n文件已保存至：\n{output_path}")
        
        except PermissionError:
            self.status_var.set("处理失败")
            messagebox.showerror("权限错误", "无法保存文件，请检查：\n1. 文件是否被其他程序打开\n2. 是否有写入权限")
        except Exception as e:
            self.status_var.set("处理失败")
            messagebox.showerror("处理错误", f"处理过程中发生错误：\n{str(e)}")
    
    def _get_output_path(self):
        """生成输出文件路径"""
        dir_name = os.path.dirname(self.file_path)
        file_name = os.path.basename(self.file_path)
        name, ext = os.path.splitext(file_name)
        output_name = f"{name}_processed{ext}"
        output_path = os.path.join(dir_name, output_name)
        return os.path.normpath(output_path)


def main():
    """程序主入口"""
    if HAS_DND:
        root = TkinterDnD.Tk()
    else:
        root = tk.Tk()
    
    style = ttk.Style()
    style.theme_use('clam')
    
    app = ExcelProcessorApp(root)
    
    root.mainloop()


if __name__ == "__main__":
    main()
