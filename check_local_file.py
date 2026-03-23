# -*- coding: utf-8 -*-
"""
检查本地test.xlsx文件
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取本地文件
local_file = r'C:\Users\12824\Desktop\dataLabel\0319\p12\test.xlsx'

try:
    # 用pandas查看数据
    df = pd.read_excel(local_file)
    print("File content:")
    print(df)
    print("\nColumn info:")
    print("Columns:", df.columns.tolist())
    print("Shape:", df.shape)
    
    # 检查第4列数据
    if len(df.columns) >= 4:
        col4_name = df.columns[3]
        print("\nColumn 4 (index 3) name:", col4_name)
        print("Column 4 values:", df.iloc[:, 3].tolist())
        print("Values > 3:", df[df.iloc[:, 3] > 3].iloc[:, 3].tolist())
    
    # 用openpyxl检查颜色
    wb = load_workbook(local_file)
    ws = wb.active
    
    print("\n" + "="*50)
    print("Checking cell colors in column 4 (D column):")
    
    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    highlighted_cells = []
    
    max_row = ws.max_row
    for row_idx in range(2, max_row + 1):  # 从第2行开始（跳过表头）
        cell = ws.cell(row=row_idx, column=4)  # 第4列
        cell_value = cell.value
        cell_fill = cell.fill
        
        # 检查颜色
        fill_color = cell_fill.start_color.rgb if cell_fill and cell_fill.start_color else None
        
        print("Row {} (Value: {}) - Fill color: {}".format(row_idx, cell_value, fill_color))
        
        if fill_color in ['FFFF0000', 'FF0000']:
            highlighted_cells.append((row_idx, cell_value))
    
    print("\nHighlighted cells (>3 should be red):", highlighted_cells)
    wb.close()
    
except Exception as e:
    print("Error: {}".format(str(e)))
    import traceback
    traceback.print_exc()
